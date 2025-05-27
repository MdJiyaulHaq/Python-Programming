import openpyxl

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("sample.xlsx")

# Displaying the names of all sheets in the workbook
print(wb.sheetnames)

# Accessing the active sheet
sheet = wb.active

# Displaying the title of the active sheet and some cell values
print(sheet.title)

# Iterating through rows and columns
for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=col)
        print(cell.value, end=" ")
print("\n")

# Accessing specific cells
print(sheet["A1"].value)
print(sheet["A2"].value)

# Accessing a specific sheet
sheet = wb["Sheet1"]
print(sheet["A1"].value)

# Accessing a specific cell
cell = sheet.cell(row=1, column=1)
print(cell.value)

# Modifying a cell
cell.value = "Hello, World!"

# Saving the workbook
wb.save("sample_modified.xlsx")

# Adding a new sheet
new_sheet = wb.create_sheet(title="NewSheet")

# Adding data to the new sheet
new_sheet["A1"] = "This is a new sheet"

# Saving the workbook again
wb.save("sample_modified.xlsx")

# Deleting a sheet
if "NewSheet" in wb.sheetnames:
    del wb["NewSheet"]

# Saving the workbook after deletion
wb.save("sample_modified.xlsx")

# Closing the workbook
wb.close()
